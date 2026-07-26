from __future__ import annotations

from collections import defaultdict
from io import BytesIO
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import select
from sqlalchemy.orm import selectinload

from ..database import SessionDep
from ..models import (
    Adviser,
    DocumentType,
    ExtractionSchemaStatus,
    ProgramAdviserAssignment,
    RequirementSlot,
    RequirementSlotItem,
    SchoolYear,
    Student,
    User,
    UserRole,
)
from .helpers import get_program_id_to_department_name_map
from .requirements import get_bulk_student_slot_statuses, list_requirement_slots

COLLEGE_NAME = "College of Computing Studies"

# ── Shared openpyxl helpers ───────────────────────────────────────────────────

_HEADER_FONT = Font(name="Arial", bold=True, size=11)
_TITLE_FONT = Font(name="Arial", bold=True, size=13)
_SUBTITLE_FONT = Font(name="Arial", bold=False, size=10)


def _write_header_row(ws, row: int, values: list[str]) -> int:
    """Write a bold header row and return the next row number."""
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = _HEADER_FONT
    return row + 1


def _write_sheet_header(ws, report_title: str, school_year_name: str) -> int:
    """Write the college title + report subtitle at the top of a sheet.

    Returns the row number after the blank separator row (ready for the
    first data header row).
    """
    ws.cell(row=1, column=1, value=COLLEGE_NAME).font = _TITLE_FONT
    ws.cell(row=2, column=1, value=f"{report_title} — {school_year_name}").font = _SUBTITLE_FONT
    return 4  # row 3 stays blank


# ── Student Report ────────────────────────────────────────────────────────────


async def build_students_xlsx(db: SessionDep) -> bytes:
    """Multi-sheet workbook — one sheet per school year.

    Columns: First Name, Last Name, Email, Student Number, Department,
    one column per required document type (True/False for verified).
    Summary at bottom: per-department student count + total.
    """
    wb = Workbook()
    # Remove default sheet — we add one per school year.
    wb.remove(wb.active)

    # Load every school year, active first.
    sy_result = await db.execute(
        select(SchoolYear).order_by(SchoolYear.is_active.desc(), SchoolYear.start_date.desc())
    )
    school_years = list(sy_result.scalars().all())

    for sy in school_years:
        await _build_student_sheet(db, wb, sy)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def _build_student_sheet(db, wb: Workbook, sy: SchoolYear) -> None:
    """Add one sheet for a single school year."""
    ws = wb.create_sheet(title=_sanitize_sheet_name(sy.name))
    next_row = _write_sheet_header(ws, "Student Report", sy.name)

    # ── Requirement slots for this school year (dynamic columns) ─
    slots = await list_requirement_slots(db, sy.id)
    slot_display_names = [
        slot.group_name or slot.description or (
            slot.items[0].document_type_name if slot.items else f"Slot {slot.display_order}"
        )
        for slot in slots
    ]

    # ── Students in this school year ──────────────────────────────────
    student_result = await db.execute(
        select(User)
        .where(User.role == UserRole.STUDENT)
        .options(
            selectinload(User.student).selectinload(Student.program_department),
        )
        .join(User.student)
        .where(Student.school_year_id == sy.id)
        .order_by(User.last_name, User.first_name)
    )
    students = list(student_result.scalars().all())

    if not students:
        return

    # ── Header row ────────────────────────────────────────────────────
    headers = ["First Name", "Last Name", "Email", "Student Number", "Department"]
    headers += slot_display_names
    next_row = _write_header_row(ws, next_row, headers)

    # ── Data rows + accumulate per-department counts ─────────────────
    student_list = [u.student for u in students if u.student is not None]
    statuses_map = await get_bulk_student_slot_statuses(db, student_list)

    dept_counts: dict[str, int] = defaultdict(int)
    for user in students:
        student = user.student
        if student is None:
            continue
        dept = student.program_department
        dept_name = dept.name if dept else "Unknown"
        dept_counts[dept_name] += 1

        slot_statuses = statuses_map.get(student.id, [])
        slot_status_map = {str(s.id): s.is_complete for s in slot_statuses}

        row = [
            user.first_name or "",
            user.last_name or "",
            user.email or "",
            student.student_number or "",
            dept_name,
        ]
        for slot in slots:
            status = slot_status_map.get(str(slot.id))
            if status is None:
                row.append("N/A")
            else:
                row.append("Yes" if status else "No")
        for col_idx, val in enumerate(row, start=1):
            ws.cell(row=next_row, column=col_idx, value=val)
        next_row += 1

    # ── Summary section ───────────────────────────────────────────────
    _write_summary_section(ws, next_row, "Department", "Student Count", dept_counts)


# ── Adviser Report ────────────────────────────────────────────────────────────


async def build_advisers_xlsx(db: SessionDep) -> bytes:
    """Multi-sheet workbook — one sheet per school year.

    Columns: First Name, Last Name, Email, Current Assignment.
    Summary at bottom: per-department adviser count + total.
    """
    wb = Workbook()
    wb.remove(wb.active)

    sy_result = await db.execute(
        select(SchoolYear).order_by(SchoolYear.is_active.desc(), SchoolYear.start_date.desc())
    )
    school_years = list(sy_result.scalars().all())

    # Pre-load the program-id → department-name map (stable across years).
    dept_name_map = await get_program_id_to_department_name_map(db)

    for sy in school_years:
        await _build_adviser_sheet(db, wb, sy, dept_name_map)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def _build_adviser_sheet(
    db,
    wb: Workbook,
    sy: SchoolYear,
    dept_name_map: dict[UUID, str],
) -> None:
    """Add one sheet for a single school year."""
    ws = wb.create_sheet(title=_sanitize_sheet_name(sy.name))
    next_row = _write_sheet_header(ws, "Adviser Report", sy.name)

    # ── Advisers with their assignments for this year ─────────────────
    adviser_result = await db.execute(
        select(User)
        .where(User.role == UserRole.ADVISER)
        .options(
            selectinload(User.adviser).selectinload(
                Adviser.program_adviser_assignments
            ).selectinload(ProgramAdviserAssignment.school_year),
        )
        .order_by(User.last_name, User.first_name)
    )
    advisers = list(adviser_result.scalars().all())

    # ── Header row ────────────────────────────────────────────────────
    next_row = _write_header_row(
        ws, next_row, ["First Name", "Last Name", "Email", "Current Assignment"]
    )

    # ── Data rows + per-department counts ────────────────────────────
    dept_counts: dict[str, int] = defaultdict(int)
    for user in advisers:
        adviser = user.adviser
        if adviser is None:
            continue

        # Collect assignments in this school year.
        yr_assignments = [
            a for a in (adviser.program_adviser_assignments or [])
            if a.school_year_id == sy.id
        ]
        assignment_count = len(yr_assignments)

        if yr_assignments:
            dept_names: list[str] = []
            for a in yr_assignments:
                dname = dept_name_map.get(a.program_id)
                if dname and dname not in dept_names:
                    dept_names.append(dname)
                    dept_counts[dname] += 1
            assignment = ", ".join(dept_names) if dept_names else f"{len(yr_assignments)} program(s)"
        else:
            assignment = "Unassigned"

        row = [
            user.first_name or "",
            user.last_name or "",
            user.email or "",
            assignment,
        ]
        for col_idx, val in enumerate(row, start=1):
            ws.cell(row=next_row, column=col_idx, value=val)
        next_row += 1

    # ── Summary section ───────────────────────────────────────────────
    _write_summary_section(ws, next_row, "Department", "Adviser Count", dept_counts)


# ── Document Requirements Report ──────────────────────────────────────────────


async def build_document_requirements_xlsx(db: SessionDep) -> bytes:
    """Multi-sheet workbook — one sheet per school year.

    Columns: Document Type, Code, Extraction Schema, Schema Version, Schema Status.
    Summary at bottom: total document types, structured-vs-classification-only counts.
    """
    wb = Workbook()
    wb.remove(wb.active)

    sy_result = await db.execute(
        select(SchoolYear).order_by(SchoolYear.is_active.desc(), SchoolYear.start_date.desc())
    )
    school_years = list(sy_result.scalars().all())

    for sy in school_years:
        await _build_document_requirements_sheet(db, wb, sy)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def _build_document_requirements_sheet(
    db,
    wb: Workbook,
    sy: SchoolYear,
) -> None:
    """Add one sheet for a single school year."""
    ws = wb.create_sheet(title=_sanitize_sheet_name(sy.name))
    next_row = _write_sheet_header(ws, "Document Requirements Report", sy.name)

    # ── Requirements for this school year via slot items ───────────────
    req_result = await db.execute(
        select(RequirementSlot)
        .where(RequirementSlot.school_year_id == sy.id)
        .options(
            selectinload(RequirementSlot.items)
            .selectinload(RequirementSlotItem.document_type),
            selectinload(RequirementSlot.items)
            .selectinload(RequirementSlotItem.extraction_schema),
        )
        .order_by(RequirementSlot.display_order)
    )
    slots = list(req_result.scalars().all())

    # ── Header row ────────────────────────────────────────────────────
    next_row = _write_header_row(
        ws, next_row,
        ["Document Type", "Code", "Extraction Schema", "Schema Version", "Schema Status"],
    )

    structured_count = 0
    classification_only_count = 0

    for slot in slots:
        for item in slot.items:
            dt = item.document_type
            schema = item.extraction_schema

            has_schema = (
                schema is not None
                and schema.status == ExtractionSchemaStatus.ACTIVE
            )

            if has_schema:
                structured_count += 1
                schema_name = schema.name
                schema_version = schema.version_label or ""
                schema_status = schema.status.value
            else:
                classification_only_count += 1
                schema_name = "—"
                schema_version = "—"
                schema_status = "Classification-only"

            row = [
                dt.name if dt else "",
                dt.code if dt else "",
                schema_name,
                schema_version,
                schema_status,
            ]
            for col_idx, val in enumerate(row, start=1):
                ws.cell(row=next_row, column=col_idx, value=val)
            next_row += 1

    # ── Summary section ───────────────────────────────────────────────
    total = structured_count + classification_only_count
    next_row += 1  # blank separator row
    ws.cell(row=next_row, column=1, value="Summary").font = _HEADER_FONT
    next_row += 1
    ws.cell(row=next_row, column=1, value="Total Document Types").font = _HEADER_FONT
    ws.cell(row=next_row, column=2, value=total)
    next_row += 1
    ws.cell(row=next_row, column=1, value="With Structured Extraction").font = _HEADER_FONT
    ws.cell(row=next_row, column=2, value=structured_count)
    next_row += 1
    ws.cell(row=next_row, column=1, value="Classification Only").font = _HEADER_FONT
    ws.cell(row=next_row, column=2, value=classification_only_count)


# ── Shared helpers ────────────────────────────────────────────────────────────


def _write_summary_section(
    ws,
    start_row: int,
    label_col1: str,
    label_col2: str,
    dept_counts: dict[str, int],
) -> None:
    """Write a per-department count table and a grand-total row."""
    next_row = start_row + 1  # blank separator row

    ws.cell(row=next_row, column=1, value="Summary").font = _HEADER_FONT
    next_row += 1

    _write_header_row(ws, next_row, [label_col1, label_col2])
    next_row += 1

    total = 0
    for dname in sorted(dept_counts):
        count = dept_counts[dname]
        ws.cell(row=next_row, column=1, value=dname)
        ws.cell(row=next_row, column=2, value=count)
        total += count
        next_row += 1

    # Grand total
    ws.cell(row=next_row, column=1, value="Total").font = _HEADER_FONT
    ws.cell(row=next_row, column=2, value=total).font = _HEADER_FONT


def _sanitize_sheet_name(name: str) -> str:
    """openpyxl sheet names are limited to 31 characters and cannot contain
    ``[ ] : * ? / \\``."""
    forbidden = {"[", "]", ":", "*", "?", "/", "\\"}
    cleaned = "".join(c for c in name if c not in forbidden)
    return cleaned[:31]
