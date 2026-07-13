from __future__ import annotations

from io import BytesIO
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import select

from ..database import SessionDep
from ..models import SchoolYear
from .admin_analytics.discovery import get_canonical_keys
from .admin_analytics.snapshot import get_extraction_analytics

COLLEGE_NAME = "College of Computing Studies"

_HEADER_FONT = Font(name="Arial", bold=True, size=11)
_TITLE_FONT = Font(name="Arial", bold=True, size=13)
_SUBTITLE_FONT = Font(name="Arial", bold=False, size=10)


def _sanitize_sheet_name(name: str) -> str:
    forbidden = {"[", "]", ":", "*", "?", "/", "\\"}
    cleaned = "".join(c for c in name if c not in forbidden)
    return cleaned[:31]


def _write_header_row(ws, row: int, values: list[str]) -> int:
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = _HEADER_FONT
    return row + 1


def _write_sheet_header(ws, report_title: str, school_year_name: str) -> int:
    ws.cell(row=1, column=1, value=COLLEGE_NAME).font = _TITLE_FONT
    ws.cell(row=2, column=1, value=f"{report_title} \u2014 {school_year_name}").font = _SUBTITLE_FONT
    return 4


def _field_summary(field: dict) -> str:
    mode = field["analytics_mode"]
    if mode == "distribution":
        dist = field.get("distribution") or []
        return f"{len(dist)} unique values"
    elif mode == "numeric_summary":
        parts = []
        if field.get("mean") is not None:
            parts.append(f"Mean: {round(field['mean'], 2)}")
        if field.get("median") is not None:
            parts.append(f"Median: {round(field['median'], 2)}")
        if field.get("min") is not None and field.get("max") is not None:
            parts.append(f"Range: {round(field['min'], 2)}\u2013{round(field['max'], 2)}")
        return ", ".join(parts) if parts else "\u2014"
    elif mode == "boolean_summary":
        t = field.get("true") or {}
        f = field.get("false") or {}
        tc = t.get("count", 0)
        fc = f.get("count", 0)
        total = tc + fc
        tp = round(tc / total * 100, 1) if total else 0
        fp = round(fc / total * 100, 1) if total else 0
        return f"True: {tc} ({tp}%), False: {fc} ({fp}%)"
    return "\u2014"


async def build_analytics_xlsx(
    db: SessionDep,
    school_year_ids: list[str],
    department_id: str | None = None,
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    sy_ids = [UUID(sid) for sid in school_year_ids]
    sy_result = await db.execute(
        select(SchoolYear)
        .where(SchoolYear.id.in_(sy_ids))
        .order_by(SchoolYear.start_date.desc())
    )
    school_years = list(sy_result.scalars().all())

    if not school_years:
        ws = wb.create_sheet(title="No Data")
        ws.cell(row=1, column=1, value="No matching school years found.")
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    dept = UUID(department_id) if department_id else None

    snapshots: list[tuple[str, dict]] = []
    for sy in school_years:
        snapshot = await get_extraction_analytics(db, sy.id, department_id=dept)
        snapshots.append((sy.name, snapshot))

    # ── Summary Sheet ──
    ws_summary = wb.create_sheet(title="Summary")
    _write_sheet_header(ws_summary, "Analytics Report", "All Selected Years")
    next_row = _write_header_row(
        ws_summary, 4,
        ["School Year", "Total Students", "Verified Submissions", "Fields Tracked", "Avg Completion %"],
    )
    for sy_name, snapshot in snapshots:
        fields = snapshot.get("fields", [])
        avg = (
            round(sum(f["insights"]["completion_rate"] for f in fields) / len(fields), 1)
            if fields
            else 0.0
        )
        ws_summary.cell(row=next_row, column=1, value=sy_name)
        ws_summary.cell(row=next_row, column=2, value=snapshot["total_students"])
        ws_summary.cell(row=next_row, column=3, value=snapshot["total_verified_submissions"])
        ws_summary.cell(row=next_row, column=4, value=len(fields))
        ws_summary.cell(row=next_row, column=5, value=avg)
        next_row += 1

    # ── One Field Analytics sheet per school year ──
    for sy_name, snapshot in snapshots:
        _build_field_analytics_sheet(wb, sy_name, snapshot)

    # ── Document Compliance sheet (all years combined) ──
    _build_compliance_sheet(wb, snapshots)

    # ── Canonical Keys sheet ──
    keys_list = await get_canonical_keys(db)
    _build_canonical_keys_sheet(wb, keys_list)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_field_analytics_sheet(wb: Workbook, sy_name: str, snapshot: dict) -> None:
    ws = wb.create_sheet(title=_sanitize_sheet_name(f"Analytics {sy_name}"))
    _write_sheet_header(ws, "Field Analytics", sy_name)

    ws.cell(row=4, column=1, value="Total Students").font = _HEADER_FONT
    ws.cell(row=4, column=2, value=snapshot["total_students"])
    ws.cell(row=5, column=1, value="Verified Submissions").font = _HEADER_FONT
    ws.cell(row=5, column=2, value=snapshot["total_verified_submissions"])

    fields = snapshot.get("fields", [])
    next_row = 7
    next_row = _write_header_row(
        ws, next_row,
        ["Group", "Label", "Canonical Key", "Type", "Mode", "Present", "Missing", "Completion %", "Summary"],
    )

    for field in fields:
        insights = field["insights"]
        ws.cell(row=next_row, column=1, value=field.get("analytics_group") or "")
        ws.cell(row=next_row, column=2, value=field["label"])
        ws.cell(row=next_row, column=3, value=field["canonical_key"])
        ws.cell(row=next_row, column=4, value=field["field_type"])
        ws.cell(row=next_row, column=5, value=field["analytics_mode"])
        ws.cell(row=next_row, column=6, value=insights["values_present"])
        ws.cell(row=next_row, column=7, value=insights["values_missing"])
        ws.cell(row=next_row, column=8, value=insights["completion_rate"])
        ws.cell(row=next_row, column=9, value=_field_summary(field))
        next_row += 1

    # ── Distribution detail tables below the main table ──
    dist_fields = [f for f in fields if f["analytics_mode"] == "distribution" and f.get("distribution")]
    if dist_fields:
        next_row += 2
        ws.cell(row=next_row, column=1, value="Distribution Details").font = _HEADER_FONT
        next_row += 2

        for field in dist_fields:
            distribution = field["distribution"] or []
            if not distribution:
                continue
            ws.cell(row=next_row, column=1, value=field["label"]).font = _HEADER_FONT
            next_row += 1
            next_row = _write_header_row(ws, next_row, ["Value", "Count", "Percentage"])
            for item in distribution:
                ws.cell(row=next_row, column=1, value=item.get("value", ""))
                ws.cell(row=next_row, column=2, value=item.get("count", 0))
                ws.cell(row=next_row, column=3, value=round(item.get("percentage", 0), 1))
                next_row += 1
            next_row += 1


def _build_compliance_sheet(wb: Workbook, snapshots: list[tuple[str, dict]]) -> None:
    all_compliance: list[tuple[str, dict]] = []
    for sy_name, snapshot in snapshots:
        for item in snapshot.get("document_compliance", []):
            all_compliance.append((sy_name, item))

    if not all_compliance:
        return

    ws = wb.create_sheet(title="Document Compliance")
    _write_sheet_header(ws, "Document Compliance", "All Selected Years")

    next_row = _write_header_row(
        ws, 4,
        ["School Year", "Document Type", "Code", "Eligible Students", "Verified", "Pending", "Missing", "Rate %"],
    )
    for sy_name, item in all_compliance:
        ws.cell(row=next_row, column=1, value=sy_name)
        ws.cell(row=next_row, column=2, value=item["document_type"])
        ws.cell(row=next_row, column=3, value=item.get("document_code", ""))
        ws.cell(row=next_row, column=4, value=item["eligible_students"])
        ws.cell(row=next_row, column=5, value=item["verified"])
        ws.cell(row=next_row, column=6, value=item["pending"])
        ws.cell(row=next_row, column=7, value=item["missing"])
        ws.cell(row=next_row, column=8, value=item["verification_rate"])
        next_row += 1


def _build_canonical_keys_sheet(wb: Workbook, keys_list: list[dict]) -> None:
    if not keys_list:
        return

    ws = wb.create_sheet(title="Canonical Keys")
    _write_sheet_header(ws, "Canonical Keys Registry", "All Time")

    next_row = _write_header_row(
        ws, 4,
        ["Canonical Key", "Label", "Field Type", "Analytics Group", "SY Count", "Document Types"],
    )
    for key in keys_list:
        ws.cell(row=next_row, column=1, value=key.get("canonical_key", ""))
        ws.cell(row=next_row, column=2, value=key.get("label", ""))
        ws.cell(row=next_row, column=3, value=key.get("field_type", ""))
        ws.cell(row=next_row, column=4, value=key.get("analytics_group") or "")
        ws.cell(row=next_row, column=5, value=key.get("school_year_count", 0))
        ws.cell(row=next_row, column=6, value=", ".join(key.get("document_types", [])))
        next_row += 1
